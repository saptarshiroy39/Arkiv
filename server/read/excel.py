import io
from openpyxl import load_workbook
from server.config import logger

def read_excel(blob, fname, **kwargs):
    parts = []
    try:
        wb = load_workbook(io.BytesIO(blob), data_only=True)
        for sheet in wb:
            parts.append(f"--- {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                vals = [str(v) if v is not None else "" for v in row]
                line = "\t".join(vals)
                if line.strip(): parts.append(line)
        return [{"text": "\n".join(parts), "page": 1}] if parts else []
    except Exception as e:
        logger.error(f"Excel error {fname}: {e}")
        return []
